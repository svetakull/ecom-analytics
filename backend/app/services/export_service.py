"""
Экспорт данных логистики в Excel и CSV.
"""
import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Столбцы выходного отчёта (ТЗ раздел 8)
LOGISTICS_COLUMNS = [
    ("seller_article", "Артикул"),
    ("nm_id", "nmId"),
    ("operation_type", "Тип операции"),
    ("warehouse", "Склад"),
    ("supply_number", "Номер поставки"),
    ("operation_date", "Дата операции"),
    ("coef_fix_start", "Дата начала фикс. коэф."),
    ("coef_fix_end", "Дата окончания фикс. коэф."),
    ("warehouse_coef", "Коэф. склада"),
    ("ktr_value", "КТР на дату операции"),
    ("irp_value", "ИРП на дату операции, %"),
    ("base_first_liter", "Базовая ставка 1-го литра, ₽"),
    ("base_per_liter", "Базовая ставка доп. литров, ₽/л"),
    ("volume_card_liters", "Объём по карточке, л"),
    ("volume_nomenclature_liters", "Объём по номенклатуре, л"),
    ("calculated_wb_volume", "Расчётный замер WB, л"),
    ("retail_price", "Розничная цена до скидки WB, ₽"),
    ("expected_logistics", "Ожидаемая логистика, ₽"),
    ("actual_logistics", "Фактическая логистика, ₽"),
    ("difference", "Разница, ₽"),
    ("operation_status", "Статус операции"),
    ("dimensions_status", "Статус габаритов"),
    ("volume_difference", "Разница объёмов, л"),
    ("ktr_needs_check", "КТР требует проверки"),
    ("tariff_missing", "Тариф склада отсутствует"),
]


def export_logistics_xlsx(operations: list[dict]) -> bytes:
    """Экспорт операций в формат Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Логистика WB"

    # Стили
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Заголовки
    for col_idx, (field, label) in enumerate(LOGISTICS_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Данные
    for row_idx, op in enumerate(operations, 2):
        for col_idx, (field, _) in enumerate(LOGISTICS_COLUMNS, 1):
            val = op.get(field, "")
            if isinstance(val, bool):
                val = "Да" if val else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

            # Цветовая индикация
            if field == "operation_status":
                if val == "Переплата":
                    cell.fill = PatternFill(start_color="FDECEA", fill_type="solid")
                elif val == "Экономия":
                    cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
            elif field == "dimensions_status":
                if val == "Занижение":
                    cell.fill = PatternFill(start_color="FDECEA", fill_type="solid")
                elif val == "Превышение":
                    cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
            elif field == "ktr_needs_check" and val == "Да":
                cell.fill = PatternFill(start_color="FFF9C4", fill_type="solid")

    # Ширины столбцов
    for col_idx, (field, label) in enumerate(LOGISTICS_COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(label) + 2, 12)

    # Закрепить заголовок
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_logistics_csv(operations: list[dict]) -> bytes:
    """Экспорт операций в формат CSV (с BOM для Excel)."""
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM для корректного открытия в Excel

    writer = csv.writer(buf, delimiter=";")

    # Заголовки
    writer.writerow([label for _, label in LOGISTICS_COLUMNS])

    # Данные
    for op in operations:
        row = []
        for field, _ in LOGISTICS_COLUMNS:
            val = op.get(field, "")
            if isinstance(val, bool):
                val = "Да" if val else ""
            row.append(val)
        writer.writerow(row)

    return buf.getvalue().encode("utf-8")
