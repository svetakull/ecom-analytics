"""
Сервис себестоимости: CRUD, resolve на дату, импорт/экспорт Excel.
"""
import io
import logging
from datetime import date, datetime
from typing import Optional

from sqlalchemy import and_, or_, case, desc
from sqlalchemy.orm import Session

from app.models.catalog import Channel, ChannelType, SKU, SKUChannel
from app.models.cost_price import (
    CostPrice, CostPriceAudit, CostPriceAuditAction, CostPriceAuditSource,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
#  RESOLVE — получить себестоимость на дату
# ═══════════════════════════════════════════════════════════════════════

def resolve_cost(
    db: Session,
    sku_id: int,
    channel_id: int,
    target_date: date,
    size: str | None = None,
) -> dict:
    """
    Вернуть актуальную себестоимость на дату.
    Приоритет: историческая запись (effective_from <= target_date) > default.
    """
    q = db.query(CostPrice).filter(
        CostPrice.sku_id == sku_id,
        CostPrice.channel_id == channel_id,
        or_(CostPrice.size == size, CostPrice.size.is_(None)),
        or_(
            and_(CostPrice.is_default == False, CostPrice.effective_from <= target_date),
            CostPrice.is_default == True,
        ),
    ).order_by(
        case((CostPrice.is_default == False, 0), else_=1),
        desc(CostPrice.effective_from),
    ).first()

    if q:
        return {
            "cost_price": float(q.cost_price),
            "fulfillment": float(q.fulfillment),
            "vat_rate": float(q.vat_rate),
            "effective_from": q.effective_from.isoformat() if q.effective_from else None,
            "is_default": q.is_default,
        }
    return {"cost_price": 0, "fulfillment": 0, "vat_rate": 0, "effective_from": None, "is_default": True}


def resolve_cogs_per_unit(
    db: Session,
    sku_id: int,
    channel_id: int,
    target_date: date,
    size: str | None = None,
) -> float:
    """Возвращает себестоимость + фулфилмент (с НДС) на единицу."""
    r = resolve_cost(db, sku_id, channel_id, target_date, size)
    base = r["cost_price"] + r["fulfillment"]
    if r["vat_rate"] > 0:
        base *= (1 + r["vat_rate"] / 100)
    return round(base, 2)


# ═══════════════════════════════════════════════════════════════════════
#  CRUD
# ═══════════════════════════════════════════════════════════════════════

def list_cost_prices(
    db: Session,
    channel_id: int | None = None,
    marketplace: str | None = None,
    article: str | None = None,
) -> list[dict]:
    """Список себестоимостей с группировкой по артикулу."""
    q = db.query(CostPrice).join(SKU, CostPrice.sku_id == SKU.id)

    if channel_id:
        q = q.filter(CostPrice.channel_id == channel_id)
    if marketplace:
        type_map = {"wb": ChannelType.WB, "ozon": ChannelType.OZON, "lamoda": ChannelType.LAMODA}
        ct = type_map.get(marketplace.lower())
        if ct:
            ch = db.query(Channel).filter(Channel.type == ct).first()
            if ch:
                q = q.filter(CostPrice.channel_id == ch.id)
    if article:
        q = q.filter(SKU.seller_article.ilike(f"%{article}%"))

    q = q.order_by(SKU.seller_article, CostPrice.is_default.desc(), CostPrice.effective_from)
    rows = q.all()

    result = []
    for r in rows:
        sku = db.query(SKU).get(r.sku_id)
        sc = db.query(SKUChannel).filter(
            SKUChannel.sku_id == r.sku_id, SKUChannel.channel_id == r.channel_id
        ).first()
        ch = db.query(Channel).get(r.channel_id)
        result.append({
            "id": r.id,
            "sku_id": r.sku_id,
            "channel_id": r.channel_id,
            "seller_article": sku.seller_article if sku else "",
            "marketplace_article": sc.mp_article if sc else "",
            "marketplace": ch.type.value if ch else "",
            "size": r.size,
            "is_default": r.is_default,
            "effective_from": r.effective_from.isoformat() if r.effective_from else None,
            "cost_price": float(r.cost_price),
            "fulfillment": float(r.fulfillment),
            "vat_rate": float(r.vat_rate),
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        })
    return result


def create_cost_price(
    db: Session,
    sku_id: int,
    channel_id: int,
    cost_price: float,
    fulfillment: float = 0,
    vat_rate: float = 0,
    size: str | None = None,
    effective_from: date | None = None,
    user_id: int | None = None,
    source: str = "ui",
) -> dict:
    """Создать запись себестоимости."""
    is_default = effective_from is None
    rec = CostPrice(
        sku_id=sku_id,
        channel_id=channel_id,
        size=size,
        is_default=is_default,
        effective_from=effective_from,
        cost_price=cost_price,
        fulfillment=fulfillment,
        vat_rate=vat_rate,
    )
    db.add(rec)
    db.flush()

    _audit(db, rec.id, user_id, CostPriceAuditAction.INSERT, None, _to_dict(rec), source)
    db.commit()
    return {"id": rec.id, "created": True}


def update_cost_price(
    db: Session,
    record_id: int,
    cost_price: float | None = None,
    fulfillment: float | None = None,
    vat_rate: float | None = None,
    user_id: int | None = None,
    source: str = "ui",
) -> dict:
    """Обновить запись."""
    rec = db.query(CostPrice).get(record_id)
    if not rec:
        raise ValueError("Record not found")

    old = _to_dict(rec)
    if cost_price is not None:
        rec.cost_price = cost_price
    if fulfillment is not None:
        rec.fulfillment = fulfillment
    if vat_rate is not None:
        rec.vat_rate = vat_rate
    rec.updated_at = datetime.utcnow()

    _audit(db, rec.id, user_id, CostPriceAuditAction.UPDATE, old, _to_dict(rec), source)
    db.commit()
    return {"id": rec.id, "updated": True}


def delete_cost_price(db: Session, record_id: int, user_id: int | None = None, source: str = "ui") -> dict:
    """Удалить историческую запись. Удаление default запрещено."""
    rec = db.query(CostPrice).get(record_id)
    if not rec:
        raise ValueError("Record not found")
    if rec.is_default:
        raise ValueError("Cannot delete default record")

    _audit(db, rec.id, user_id, CostPriceAuditAction.DELETE, _to_dict(rec), None, source)
    db.delete(rec)
    db.commit()
    return {"id": record_id, "deleted": True}


# ═══════════════════════════════════════════════════════════════════════
#  BATCH UPSERT
# ═══════════════════════════════════════════════════════════════════════

def batch_upsert(
    db: Session,
    items: list[dict],
    mode: str = "update",
    user_id: int | None = None,
) -> dict:
    """
    Массовый upsert.
    mode='overwrite': удалить существующие записи по артикулам, вставить новые.
    mode='update': upsert только непустых полей.
    """
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for i, item in enumerate(items):
        try:
            sku = db.query(SKU).filter(SKU.seller_article == item.get("seller_article")).first()
            if not sku:
                errors.append({"row": i + 1, "field": "seller_article", "message": "SKU not found"})
                skipped += 1
                continue

            marketplace = item.get("marketplace", "wb").lower()
            type_map = {"wb": ChannelType.WB, "ozon": ChannelType.OZON, "lamoda": ChannelType.LAMODA}
            ct = type_map.get(marketplace)
            ch = db.query(Channel).filter(Channel.type == ct).first() if ct else None
            if not ch:
                errors.append({"row": i + 1, "field": "marketplace", "message": f"Channel not found: {marketplace}"})
                skipped += 1
                continue

            effective_from = item.get("effective_from")
            is_default = effective_from is None
            size = item.get("size")

            # Найти существующую запись
            existing = db.query(CostPrice).filter(
                CostPrice.sku_id == sku.id,
                CostPrice.channel_id == ch.id,
                CostPrice.size == size if size else CostPrice.size.is_(None),
                CostPrice.effective_from == effective_from if effective_from else CostPrice.is_default == True,
            ).first()

            if mode == "overwrite" and existing:
                old = _to_dict(existing)
                existing.cost_price = item["cost_price"]
                existing.fulfillment = item.get("fulfillment", 0)
                existing.vat_rate = item.get("vat_rate", 0)
                existing.updated_at = datetime.utcnow()
                _audit(db, existing.id, user_id, CostPriceAuditAction.UPDATE, old, _to_dict(existing), "excel_import")
                updated += 1
            elif existing and mode == "update":
                old = _to_dict(existing)
                changed = False
                if "cost_price" in item and item["cost_price"] is not None:
                    existing.cost_price = item["cost_price"]
                    changed = True
                if "fulfillment" in item and item["fulfillment"] is not None:
                    existing.fulfillment = item["fulfillment"]
                    changed = True
                if "vat_rate" in item and item["vat_rate"] is not None:
                    existing.vat_rate = item["vat_rate"]
                    changed = True
                if changed:
                    existing.updated_at = datetime.utcnow()
                    _audit(db, existing.id, user_id, CostPriceAuditAction.UPDATE, old, _to_dict(existing), "excel_import")
                    updated += 1
                else:
                    skipped += 1
            else:
                rec = CostPrice(
                    sku_id=sku.id,
                    channel_id=ch.id,
                    size=size,
                    is_default=is_default,
                    effective_from=effective_from,
                    cost_price=item["cost_price"],
                    fulfillment=item.get("fulfillment", 0),
                    vat_rate=item.get("vat_rate", 0),
                )
                db.add(rec)
                db.flush()
                _audit(db, rec.id, user_id, CostPriceAuditAction.INSERT, None, _to_dict(rec), "excel_import")
                created += 1

        except Exception as e:
            errors.append({"row": i + 1, "field": "", "message": str(e)})
            skipped += 1

    db.commit()
    return {"processed": len(items), "created": created, "updated": updated, "skipped": skipped, "errors": errors}


# ═══════════════════════════════════════════════════════════════════════
#  IMPORT / EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════

def export_excel(db: Session, channel_id: int | None = None, marketplace: str | None = None) -> bytes:
    """Экспорт себестоимостей в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = list_cost_prices(db, channel_id=channel_id, marketplace=marketplace)

    # Группируем по (seller_article, marketplace_article, marketplace, size)
    from collections import defaultdict
    grouped: dict[tuple, dict] = {}
    all_dates: set[str] = set()

    for r in rows:
        key = (r["seller_article"], r["marketplace_article"], r["marketplace"], r["size"] or "")
        if key not in grouped:
            grouped[key] = {"default": None, "history": {}}
        if r["is_default"]:
            grouped[key]["default"] = r
        elif r["effective_from"]:
            grouped[key]["history"][r["effective_from"]] = r
            all_dates.add(r["effective_from"])

    sorted_dates = sorted(all_dates)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Себестоимость"

    # Заголовки
    headers = ["Артикул продавца", "Артикул МП", "Маркетплейс", "Размер",
               "Себестоим. (default)", "Фулфилмент (default)", "НДС (default)"]
    for dt in sorted_dates:
        headers.extend([f"Себестоим. {dt}", f"Фулфилмент {dt}", f"НДС {dt}"])

    hfill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill

    row_num = 2
    for (sa, ma, mp, sz), data in sorted(grouped.items()):
        ws.cell(row_num, 1, sa)
        ws.cell(row_num, 2, ma)
        ws.cell(row_num, 3, mp)
        ws.cell(row_num, 4, sz or "")

        d = data["default"]
        if d:
            ws.cell(row_num, 5, d["cost_price"])
            ws.cell(row_num, 6, d["fulfillment"])
            ws.cell(row_num, 7, d["vat_rate"])

        for i, dt in enumerate(sorted_dates):
            h = data["history"].get(dt)
            if h:
                ws.cell(row_num, 8 + i * 3, h["cost_price"])
                ws.cell(row_num, 9 + i * 3, h["fulfillment"])
                ws.cell(row_num, 10 + i * 3, h["vat_rate"])

        row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_excel(
    db: Session,
    file_bytes: bytes,
    mode: str = "update",
    user_id: int | None = None,
) -> dict:
    """Импорт себестоимостей из Excel."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[1]]
    items = []
    errors = []

    for row_num in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_num, column=c).value for c in range(1, len(headers) + 1)]

        sa = str(cells[0] or "").strip()
        ma = str(cells[1] or "").strip()
        mp = str(cells[2] or "").strip().lower()
        sz = str(cells[3] or "").strip() or None

        if not sa or not ma or not mp:
            errors.append({"row": row_num, "field": "A-C", "message": "Обязательные поля не заполнены"})
            continue

        # Default (колонки E, F, G)
        try:
            default_cost = float(cells[4]) if cells[4] is not None else None
            default_ff = float(cells[5]) if cells[5] is not None else 0
            default_vat = float(cells[6]) if cells[6] is not None else 0
        except (ValueError, TypeError):
            errors.append({"row": row_num, "field": "E-G", "message": "Некорректные числовые значения"})
            continue

        if default_cost is None:
            errors.append({"row": row_num, "field": "E", "message": "Себестоимость по умолчанию обязательна"})
            continue

        items.append({
            "seller_article": sa,
            "marketplace_article": ma,
            "marketplace": mp,
            "size": sz,
            "cost_price": default_cost,
            "fulfillment": default_ff,
            "vat_rate": default_vat,
            "effective_from": None,
        })

        # Динамические столбцы (H+) — тройками
        col_idx = 7
        while col_idx + 2 < len(headers):
            h = headers[col_idx]
            # Ищем дату в заголовке
            dt_str = h.replace("Себестоим. ", "").replace("Себестоим.", "").strip()
            try:
                from datetime import datetime as dt_mod
                eff_date = dt_mod.strptime(dt_str, "%Y-%m-%d").date()
            except ValueError:
                col_idx += 3
                continue

            cost_val = cells[col_idx] if col_idx < len(cells) else None
            ff_val = cells[col_idx + 1] if col_idx + 1 < len(cells) else None
            vat_val = cells[col_idx + 2] if col_idx + 2 < len(cells) else None

            if cost_val is not None:
                try:
                    items.append({
                        "seller_article": sa,
                        "marketplace_article": ma,
                        "marketplace": mp,
                        "size": sz,
                        "cost_price": float(cost_val),
                        "fulfillment": float(ff_val) if ff_val is not None else 0,
                        "vat_rate": float(vat_val) if vat_val is not None else 0,
                        "effective_from": eff_date,
                    })
                except (ValueError, TypeError):
                    pass

            col_idx += 3

    result = batch_upsert(db, items, mode=mode, user_id=user_id)
    result["errors"].extend(errors)
    return result


# ═══════════════════════════════════════════════════════════════════════
#  УТИЛИТЫ
# ═══════════════════════════════════════════════════════════════════════

def _to_dict(rec: CostPrice) -> dict:
    return {
        "cost_price": float(rec.cost_price),
        "fulfillment": float(rec.fulfillment),
        "vat_rate": float(rec.vat_rate),
        "effective_from": rec.effective_from.isoformat() if rec.effective_from else None,
        "is_default": rec.is_default,
    }


def _audit(
    db: Session,
    cost_price_id: int,
    user_id: int | None,
    action: CostPriceAuditAction,
    old_values: dict | None,
    new_values: dict | None,
    source: str = "ui",
):
    src_map = {"ui": CostPriceAuditSource.UI, "excel_import": CostPriceAuditSource.EXCEL_IMPORT, "api": CostPriceAuditSource.API}
    db.add(CostPriceAudit(
        cost_price_id=cost_price_id,
        user_id=user_id,
        action=action,
        old_values=old_values,
        new_values=new_values,
        source=src_map.get(source, CostPriceAuditSource.UI),
    ))
